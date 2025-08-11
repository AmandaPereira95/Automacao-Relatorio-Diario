import os
import pandas as pd
import matplotlib.pyplot as plt
import openpyxl as xl
import numpy as np
import yagmail
import json
from datetime import datetime
from fpdf import FPDF
from dotenv import load_dotenv

class RelatorioPDF(FPDF):
    """Classe responsável por montar o layout do relatório PDF."""

    def header(self):
        self.set_font('Arial', 'B', 14)
        self.cell(0, 10, 'Relatório Diário de Vendas', ln=True, align='C')
        self.set_font('Arial', '', 10)
        self.cell(0, 10, f'Gerado em: {datetime.now().strftime("%d/%m/%Y %H:%M")}', ln=True, align='C')
        self.ln(10)

    def add_estatisticas(self, estatisticas):
        self.set_font('Arial', 'B', 12)
        self.cell(0, 10, 'Resumo Executivo:', ln=True)
        self.set_font('Arial', '', 11)
        for i in range(len(estatisticas["Métrica"])):
            linha = f"{estatisticas['Métrica'][i]}: R$ {estatisticas['Valor (R$)'][i]:,.2f}" \
                if isinstance(estatisticas['Valor (R$)'][i], (float, int)) else \
                f"{estatisticas['Métrica'][i]}: {estatisticas['Valor (R$)'][i]}"
            self.cell(0, 8, linha, ln=True)
        self.ln(10)

    def add_grafico(self, imagem_path, titulo):
        if os.path.exists(imagem_path):
            self.set_font('Arial', 'B', 12)
            self.cell(0, 10, titulo, ln=True)
            self.image(imagem_path, w=170)
            self.ln(10)

class RelatorioVendasRPA:
    """Classe principal que executa todas as etapas do relatório automatizado."""

    def __init__(self, config_path):
        load_dotenv()
        self.config = self.ler_arquivo_config(config_path)
        self.config["email"]["sender"] = os.getenv("EMAIL")
        self.config["email"]["password"] = os.getenv("EMAIL_SENHA")

    def ler_arquivo_config(self, caminho):
        """Carrega as configurações do arquivo settings.json."""
        with open(caminho, encoding="utf-8") as f:
            return json.load(f)

    def leitura_planilha_base(self):
        """Lê a planilha Excel com os dados brutos."""
        df = pd.read_excel(self.config['input_path'], 'Sheet1')
        return df

    def consolidar_dados(self, df):
        """Gera os agrupamentos por vendedor, região e total geral."""
        por_vendedor = df.groupby('Vendedor').agg(
            total_vendas=('Valor da Venda (R$)', 'sum'),
            quantidade_vendas=('Valor da Venda (R$)', 'count')
        ).reset_index()

        por_regiao = df.groupby('Região').agg(
            total_vendas=('Valor da Venda (R$)', 'sum'),
            quantidade_vendas=('Valor da Venda (R$)', 'count')
        ).reset_index()

        total_geral = float(np.round(df['Valor da Venda (R$)'].sum(), 2))

        estatisticas = {
            'Total de Vendas': total_geral,
            'Média por Venda': float(np.round(df['Valor da Venda (R$)'].mean(), 2)),
            'Maior Venda': float(np.round(df['Valor da Venda (R$)'].max(), 2)),
            'Menor Venda': float(np.round(df['Valor da Venda (R$)'].min(), 2)),
            'Quantidade de Vendas': len(df)
        }

        df_estatisticas = pd.DataFrame({
            "Métrica": list(estatisticas.keys()),
            "Valor (R$)": list(estatisticas.values())
        })

        return por_vendedor, por_regiao, total_geral, df_estatisticas

    def gerar_graficos(self, df_vendedores, df_regioes):
        """Gera e salva os gráficos de vendas por vendedor e por região."""
        df_vendedores.plot(kind='bar', x='Vendedor', y='total_vendas')
        plt.title('Vendas por Vendedor')
        plt.tight_layout()
        plt.savefig('./output/vendas_por_vendedor.png')
        plt.close()

        df_regioes.plot(kind='bar', x='Região', y='total_vendas')
        plt.title('Vendas por Região')
        plt.tight_layout()
        plt.savefig('./output/vendas_por_regiao.png')
        plt.close()

    def gerar_planilha_consolidada(self, df_estatisticas, df_vendedores, df_regioes):
        """Salva os dados consolidados em um Excel com abas distintas."""
        with pd.ExcelWriter(self.config['output_excel'], engine='openpyxl') as writer:
            df_estatisticas.to_excel(writer, sheet_name='Resumo', index=False)
            df_vendedores.to_excel(writer, sheet_name='VendasPorVendedor', index=False)
            df_regioes.to_excel(writer, sheet_name='VendasPorRegiao', index=False)

    def gerar_pdf(self, df_estatisticas):
        """Gera o relatório PDF com os gráficos e estatísticas."""
        estatisticas_dict = df_estatisticas.to_dict(orient="list")
        pdf = RelatorioPDF()
        pdf.add_page()
        pdf.add_estatisticas(estatisticas_dict)
        pdf.add_grafico("./output/vendas_por_vendedor.png", "Vendas por Vendedor")
        pdf.add_grafico("./output/vendas_por_regiao.png", "Vendas por Região")
        pdf.output(self.config['output_pdf'])

    def enviar_email(self, total_vendas):
        """Envia o e-mail com o Excel e PDF gerados em anexo."""
        email_cfg = self.config['email']
        email_cfg["total_vendas"] = f"R$ {total_vendas:,.2f}"

        try:
            yag = yagmail.SMTP(user=email_cfg["sender"], password=email_cfg["password"])
            yag.send(
                to=email_cfg["receiver"],
                subject=email_cfg["subject"],
                contents=f"""
                Olá,

                Segue em anexo o relatório diário de vendas gerado automaticamente pelo robô.

                📊 Total de Vendas: {email_cfg['total_vendas']}

                Atenciosamente,
                Bot RPA de Relatórios
                """,
                attachments=[
                    self.config['output_excel'],
                    self.config['output_pdf']
                ]
            )
            print("\n✅ E-mail enviado com sucesso!")

        except Exception as e:
            print(f"\n❌ Erro ao enviar e-mail: {e}")

    def executar(self):
        """Executa todas as etapas da automação."""
        print("Iniciando execução do robô de relatório...")
        df_base = self.leitura_planilha_base()
        df_vendedores, df_regioes, total, df_estatisticas = self.consolidar_dados(df_base)
        self.gerar_graficos(df_vendedores, df_regioes)
        self.gerar_planilha_consolidada(df_estatisticas, df_vendedores, df_regioes)
        self.gerar_pdf(df_estatisticas)
        self.enviar_email(total)
        print("\n🚀 Processo finalizado com sucesso.")

# Execução
if __name__ == "__main__":
    rob = RelatorioVendasRPA("./config/settings.json")
    rob.executar()